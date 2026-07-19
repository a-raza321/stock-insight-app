import itertools
import re
import warnings
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st

# ==========================================================================
# BLOCK 1: PAGE CONFIGURATION AND CONFIG CONSTANTS
# ==========================================================================
st.set_page_config(
    page_title="Financial Reconciliation Tool",
    page_icon="🗂️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls", "pdf"}
FIXED_FIELDS = ["Amount", "Amount 2", "Date", "Reference", "Description"]
BLANK_OPTION = "Select"

# ==========================================================================
# BLOCK 2: SESSION STATE INITIALIZATION & RESET UTILITIES
# ==========================================================================
DEFAULT_STATE = {
    "processed": False,
    "df1": None,
    "df2": None,
    "file1_bytes": None,
    "file1_name": None,
    "file2_bytes": None,
    "file2_name": None,
    "uploader1_key": 0,
    "uploader2_key": 0,
    "mapping_result": None,
    "reconciled": False,
    "recon_results": None,
}

# Ensure all default keys are successfully pre-populated to avoid KeyErrors
for k, v in DEFAULT_STATE.items():
    if k not in st.session_state:
        st.session_state[k] = v


def clear_mapping_widget_state():
    """Removes previously stored user dropdown choices so fresh auto-detected
    defaults are calculated cleanly upon loading any new statements."""
    for f in FIXED_FIELDS:
        st.session_state.pop(f"map1_{f}", None)
        st.session_state.pop(f"map2_{f}", None)


def reset_app():
    """Wipes all processing variables and increments file uploader keys
    to completely reset the file inputs back to empty states."""
    st.session_state.processed = False
    st.session_state.df1 = None
    st.session_state.df2 = None
    st.session_state.file1_bytes = None
    st.session_state.file1_name = None
    st.session_state.file2_bytes = None
    st.session_state.file2_name = None
    st.session_state.mapping_result = None
    st.session_state.reconciled = False
    st.session_state.recon_results = None
    st.session_state.uploader1_key += 1
    st.session_state.uploader2_key += 1
    clear_mapping_widget_state()


def delete_file(slot: int):
    """Removes a single statement instance, cleaning dependencies and letting
    the user pick a brand-new file for that slot."""
    if slot == 1:
        st.session_state.file1_bytes = None
        st.session_state.file1_name = None
        st.session_state.uploader1_key += 1
    else:
        st.session_state.file2_bytes = None
        st.session_state.file2_name = None
        st.session_state.uploader2_key += 1

    # Invalidate current reconciliation outputs since the input sheets changed
    st.session_state.processed = False
    st.session_state.df1 = None
    st.session_state.df2 = None
    st.session_state.mapping_result = None
    st.session_state.reconciled = False
    st.session_state.recon_results = None
    clear_mapping_widget_state()


# ==========================================================================
# BLOCK 3: BEAUTIFIED MODERN CSS DESIGN TOKENS
# ==========================================================================
st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(180deg, #f7f9fc 0%, #eef1f8 100%);
    }
    .app-header {
        text-align: center;
        padding: 1.6rem 1rem 1.2rem 1rem;
        margin-bottom: 1.2rem;
        border-radius: 18px;
        background: linear-gradient(120deg, #1e3a8a 0%, #3b82f6 60%, #60a5fa 100%);
        color: white;
        box-shadow: 0 10px 30px rgba(59, 130, 246, 0.25);
    }
    .app-header h1 {
        margin: 0;
        font-size: 2.1rem;
        font-weight: 800;
        letter-spacing: -0.5px;
        color: white !important;
    }
    .app-header p {
        margin: 0.35rem 0 0 0;
        font-size: 1.0rem;
        opacity: 0.92;
        color: #f0fdf4 !important;
    }
    /* Balanced Typography & Readability */
    .stApp, .stApp p, .stApp span, .stApp label, .stApp li,
    div[data-testid="stMarkdownContainer"] p,
    div[data-testid="stCaptionContainer"],
    div[data-testid="stCaptionContainer"] p,
    div[data-testid="stWidgetLabel"] p,
    div[data-testid="stExpander"] summary,
    div[data-testid="stExpander"] p,
    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricValue"] {
        color: #1f2937 !important;
    }
    div[data-testid="stCaptionContainer"] p, .stCaption {
        color: #4b5563 !important;
    }
    .upload-label {
        color: #111827;
        font-size: 1.05rem;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }
    div[data-testid="stFileUploader"] {
        margin-top: 0rem;
    }
    .file-ok {
        background: #ecfdf5;
        border: 1px solid #a7f3d0;
        color: #065f46;
        padding: 0.6rem 0.9rem;
        border-radius: 10px;
        font-size: 0.9rem;
    }
    .file-err {
        background: #fef2f2;
        border: 1px solid #fecaca;
        color: #991b1b;
        padding: 0.55rem 0.8rem;
        border-radius: 10px;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }
    .file-ext-badge {
        display: inline-block;
        background: #eef2ff;
        color: #4338ca;
        font-weight: 700;
        font-size: 0.72rem;
        letter-spacing: 0.5px;
        padding: 0.15rem 0.5rem;
        border-radius: 6px;
        margin-left: 0.5rem;
        text-transform: uppercase;
    }
    .section-title {
        font-size: 1.3rem;
        font-weight: 800;
        color: #1e3a8a;
        margin: 1.6rem 0 0.6rem 0;
    }
    /* Premium dark navy background for the uploader dropzone */
    div[data-testid="stFileUploaderDropzone"] {
        border-radius: 12px !important;
        border: 2px dashed #60a5fa !important;
        background: #1e3a8a !important;
        padding: 0.4rem 0.7rem !important;
        min-height: 0 !important;
    }
    div[data-testid="stFileUploaderDropzoneInstructions"] {
        padding: 0.2rem 0 !important;
    }
    div[data-testid="stFileUploaderDropzoneInstructions"] svg {
        height: 1.1rem !important;
        width: 1.1rem !important;
        fill: #ffffff !important;
    }
    /* Styled file uploader instructions cleanly to white */
    div[data-testid="stFileUploaderDropzoneInstructions"] span {
        font-size: 0.78rem !important;
        line-height: 1.1rem !important;
        color: #ffffff !important;
    }
    div[data-testid="stFileUploaderDropzoneInstructions"] small {
        font-size: 0.68rem !important;
        color: #e2e8f0 !important;
    }
    div[data-testid="stFileUploaderDropzone"] button {
        padding: 0.25rem 0.7rem !important;
        font-size: 0.78rem !important;
        min-height: 0 !important;
    }
    div[data-testid="stFileUploader"] section {
        min-height: 0 !important;
    }
    /* Action Buttons Design */
    div[data-testid="stButton"] button[kind="primary"] {
        border-radius: 12px;
        font-weight: 700;
        font-size: 1.0rem;
        background: #1e3a8a !important;
        color: white !important;
        border: none;
        box-shadow: 0 8px 20px rgba(30, 58, 138, 0.3);
        transition: transform 0.15s ease;
        padding: 0.6rem 0;
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        background: #1d4ed8 !important;
        transform: translateY(-2px);
    }
    div[data-testid="stButton"] button[kind="primary"]:disabled {
        background: #e2e8f0 !important;
        color: #94a3b8 !important;
    }
    /* Secondary/Reset actions */
    div[data-testid="stButton"] button[kind="secondary"] {
        border-radius: 10px;
        font-weight: 700;
        border: 1.5px solid #ef4444;
        color: #ef4444 !important;
        background: white;
    }
    div[data-testid="stButton"] button[kind="secondary"]:hover {
        background: #fef2f2;
    }
    /* High-contrast styling for the Download as Excel button to fix black-on-black visibility */
    div[data-testid="stDownloadButton"] button {
        background-color: #10b981 !important; /* Emerald green background */
        color: #ffffff !important;            /* Pure white text */
        border-radius: 12px !important;
        font-weight: 700 !important;
        font-size: 1.0rem !important;
        border: none !important;
        box-shadow: 0 8px 20px rgba(16, 185, 129, 0.3) !important;
        transition: transform 0.15s ease, background-color 0.15s ease !important;
        padding: 0.6rem 0 !important;
        display: inline-flex !important;
        justify-content: center !important;
        align-items: center !important;
        width: 100% !important;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background-color: #059669 !important; /* Richer forest green on hover */
        color: #ffffff !important;            /* Keeps text white */
        transform: translateY(-2px) !important;
    }
    /* Modern Column Mapping Layout */
    .mapping-card {
        background: white;
        border-radius: 12px;
        padding: 1rem 1.4rem;
        margin-bottom: 0.8rem;
        border: 1px solid #e2e8f0;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    }
    .mapping-grid-header {
        font-weight: 800;
        color: #1e3a8a;
        font-size: 1rem;
        padding-bottom: 0.6rem;
        border-bottom: 2px solid #cbd5e1;
        margin-bottom: 0.8rem;
    }
    .badge-full {
        display:inline-block; background:#ecfdf5; color:#065f46; border:1px solid #a7f3d0;
        padding:0.2rem 0.7rem; border-radius:999px; font-weight:700; font-size:0.78rem;
    }
    .badge-partial {
        display:inline-block; background:#fffbeb; color:#92400e; border:1px solid #fde68a;
        padding:0.2rem 0.7rem; border-radius:999px; font-weight:700; font-size:0.78rem;
    }
    .badge-split {
        display:inline-block; background:#eff6ff; color:#1e40af; border:1px solid #bfdbfe;
        padding:0.2rem 0.7rem; border-radius:999px; font-weight:700; font-size:0.78rem;
    }
    .badge-unmatched {
        display:inline-block; background:#fef2f2; color:#991b1b; border:1px solid #fecaca;
        padding:0.2rem 0.7rem; border-radius:999px; font-weight:700; font-size:0.78rem;
    }
    .recon-summary-card {
        background:white; border:1px solid #e2e8f0; border-radius:14px; padding:0.9rem 1.1rem;
        box-shadow: 0 4px 14px rgba(20,20,43,0.03); text-align:center;
    }
    .recon-summary-card .num { font-size:1.6rem; font-weight:800; color:#1e3a8a; }
    .recon-summary-card .lbl { font-size:0.8rem; color:#4b5563; font-weight:600; }

    /* Background Balance Check Display */
    .balance-card {
        background: #f0fdf4;
        border: 2px solid #bbf7d0;
        border-radius: 12px;
        padding: 1.2rem;
        margin-bottom: 1.2rem;
    }
    .balance-card-error {
        background: #fef2f2;
        border: 2px solid #fecaca;
        border-radius: 12px;
        padding: 1.2rem;
        margin-bottom: 1.2rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ==========================================================================
# BLOCK 4: FILE LOADER & SHEET PARSING ENGINE
# ==========================================================================
def get_extension(filename: str) -> str:
    """Helper to safely retrieve lowercase string file extension."""
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _looks_numeric(value: str) -> bool:
    """Detect if a raw cell string conforms to float/integer shapes."""
    v = value.strip().replace(",", "")
    if not v:
        return False
    v2 = v.replace(".", "", 1).replace("-", "", 1)
    return v2.isdigit()


def detect_header_row(raw: pd.DataFrame, max_scan: int = 25) -> int:
    """
    Looks at raw sheets to find the index of the main header row, bypassing
    blank lines, metadata headers, or title blocks.
    """
    best_idx = 0
    best_score = -1.0
    n_check = min(max_scan, len(raw))

    for i in range(n_check):
        row = raw.iloc[i]
        non_null = row.notna().sum()
        if non_null == 0:
            continue

        ratio_filled = non_null / len(row)
        str_count = sum(1 for val in row if pd.notna(val) and not _looks_numeric(str(val)))
        ratio_text = str_count / non_null
        values = row.dropna().astype(str).str.strip()
        ratio_unique = (len(values.unique()) / len(values)) if len(values) else 0

        transition_bonus = 0.0
        if i + 1 < len(raw):
            next_row = raw.iloc[i + 1]
            next_non_null = next_row.notna().sum()
            if next_non_null:
                next_numeric = sum(1 for v in next_row if pd.notna(v) and _looks_numeric(str(v)))
                if next_numeric / next_non_null > ratio_text:
                    transition_bonus = 0.15

        score = (ratio_filled * 0.35) + (ratio_text * 0.35) + (ratio_unique * 0.15) + transition_bonus
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


def clean_header_labels(values):
    """Frees sheet labels from duplicates and nan values."""
    seen = {}
    labels = []
    for j, v in enumerate(values):
        label = str(v).strip() if pd.notna(v) else ""
        if not label or label.lower() == "nan":
            label = f"Column_{j + 1}"
        if label in seen:
            seen[label] += 1
            label = f"{label}_{seen[label]}"
        else:
            seen[label] = 0
        labels.append(label)
    return labels


def extract_pdf_rows(file_bytes):
    """Processes tabular records from standard PDF files using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        st.error("pdfplumber is required to extract PDF transactions. Install it via pip.")
        return None

    import io
    rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                rows.extend(table)
    return rows if rows else None


def load_dataframe(file_bytes, filename):
    """
    Parses and sanitizes CSV, XLS, XLSX, and PDF attachments.
    Detects dynamic offsets to normalize the top columns.
    Uses exact physical spreadsheet index mapping.
    """
    import io
    ext = get_extension(filename)
    if ext not in ALLOWED_EXTENSIONS:
        return None, f"Unsupported file extension: .{ext}"

    try:
        if ext == "csv":
            raw = pd.read_csv(io.BytesIO(file_bytes), header=None, dtype=str, skip_blank_lines=False)
        elif ext in ("xlsx", "xls"):
            raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
        elif ext == "pdf":
            table_rows = extract_pdf_rows(file_bytes)
            if not table_rows:
                return None, "No readable table structure detected inside PDF."
            width = max(len(r) for r in table_rows)
            padded = [r + [None] * (width - len(r)) for r in table_rows]
            raw = pd.DataFrame(padded)
        else:
            return None, f"Unsupported file extension: .{ext}"
    except Exception as e:
        return None, f"Error parsing spreadsheet format: {e}"

    if raw is None or raw.empty:
        return None, "The uploaded file contains no data rows."

    # Keep track of original raw index as physical spreadsheet row number (1-based)
    raw['original_row_num'] = raw.index + 1

    # Drop blank lines but subset only data columns (not original_row_num)
    data_cols = [c for c in raw.columns if c != 'original_row_num']
    raw = raw.dropna(subset=data_cols, how="all").reset_index(drop=True)
    if raw.empty:
        return None, "The uploaded statement lacks parsed rows."

    # Exclude original_row_num for header row detection
    raw_data_only = raw[data_cols]
    header_idx = detect_header_row(raw_data_only)
    header_values = raw_data_only.iloc[header_idx].tolist()
    columns = clean_header_labels(header_values)

    # Slice data
    data = raw.iloc[header_idx + 1:].reset_index(drop=True)
    original_row_nums = data['original_row_num'].tolist()

    # Reconstruct data
    data_final = data[data_cols].copy()
    data_final.columns = columns
    data_final['original_row_num'] = original_row_nums

    return data_final, None


# ==========================================================================
# BLOCK 5: TRANSACTION TYPE DETECTION & DATA SANITIZATION HELPERS
# ==========================================================================
FIELD_SIGNALS = {
    "Amount": ["payment amount", "amount", "net amount", "total", "value", "debit", "credit", "paid"],
    "Date": ["transaction date", "value date", "posting date", "date"],
    "Reference": ["invoice", "reference", "ref no", "ref", "cheque", "txn id", "transaction id", "receipt"],
    "Description": ["description", "narration", "particulars", "details", "memo", "remarks", "type"],
}

# Standardize transaction categories exactly to title case ("Invoice", "Payment", etc.)
KEYWORDS_MAP = {
    # INVOICE
    "sales invoice": "Invoice",
    "tax invoice": "Invoice",
    "sales inv": "Invoice",
    "invoice": "Invoice",
    "charge": "Invoice",
    "inv": "Invoice",

    # PAYMENT (CSH keyword mapped dynamically as Payment)
    "cash receipt": "Payment",
    "direct deposit": "Payment",
    "payment": "Payment",
    "receipt": "Payment",
    "cash": "Payment",
    "eft": "Payment",
    "csh": "Payment",

    # CREDIT NOTE (Including CRE and Credit Memo matching)
    "credit note": "Credit Note",
    "credit_note": "Credit Note",
    "correction": "Credit Note",
    "crn": "Credit Note",
    "cre": "Credit Note",
    "credit memo": "Credit Note",
    "credit_memo": "Credit Note",

    # ADJUSTMENT
    "adjustment": "Adjustment",
    "reversal": "Adjustment",
    "adj": "Adjustment"
}

# Sort keywords by length descending to match most descriptive phrases first
SORTED_KEYWORDS = sorted(KEYWORDS_MAP.keys(), key=len, reverse=True)


def auto_detect_column(columns, field):
    """
    Resolves matching index based on prioritized search tokens in columns list.
    """
    lowered = {c: c.lower().strip() for c in columns}
    for signal in FIELD_SIGNALS[field]:
        for col, low in lowered.items():
            if low == signal:
                return col
    for signal in FIELD_SIGNALS[field]:
        for col, low in lowered.items():
            if signal in low:
                return col
    return None


def is_blank_value(val):
    """Check if raw cell content is empty or represents a missing ledger balance."""
    if pd.isna(val):
        return True
    s = str(val).strip()
    if s == "" or s.lower() in ("nan", "none", "—", "-"):
        return True
    return False


def clean_cell_text(val):
    """Removes standard pandas formatting fillers so cell outputs look clean."""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "n/a", "none", "unknown", "—", "-"):
        return ""
    return s


def clean_amount_display(val):
    """Formats float numeric tags directly."""
    try:
        v = float(val)
        if pd.isna(v) or v == 0.0:
            return None
        return v
    except (ValueError, TypeError):
        return None


def clean_dataframe_strings(df):
    """
    Wipes pandas metadata text items like nan, n/a, unknown, none and preserves
    raw float columns intact so Streamlit's sorting engines execute normally.
    """
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            continue
        df[col] = df[col].apply(lambda x: "" if pd.isna(x) or str(x).lower().strip() in (
            "nan", "n/a", "none", "unknown", "—", "-") else str(x))
    return df


def parse_amount(val):
    """
    Extracts absolute float magnitude. Eliminates currency tags,
    commas, minus signs, and parenthetical notations (e.g. (100) -> 100.0).
    """
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return 0.0

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if "-" in s:
        neg = True

    if s.upper().endswith("CR") or s.upper().endswith("DR"):
        s = s[:-2].strip()

    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", ".", "-."):
        return 0.0
    try:
        return abs(float(s))
    except ValueError:
        return 0.0


def word_contains_kw(val, kw):
    """
    Safely matches standing keywords with custom boundaries, ensuring letters/numbers
    are aligned but prevents overlap captures (e.g., matching "eft" inside "leftover").
    """
    pattern = rf"(?<![a-z]){re.escape(kw)}(?![a-z])"
    return bool(re.search(pattern, val))


def detect_row_category(raw_row):
    """
    Scans explicit transaction category columns first (Type, Document Type, Category, etc.) to
    get standard clean accounting transaction parent categories. Falls back to description/reference
    or global string matching only if no explicit columns match.
    """
    explicit_headers = [
        "document type", "document_type", "documenttype",
        "transaction type", "transaction_type", "transactiontype", "transaction types",
        "transection type", "transection_type", "transectiontype", "transection types",
        "trasection type", "trasection_type", "trasectiontype", "trasection types",
        "type", "category", "class"
    ]

    fallback_headers = [
        "description", "reference", "ref"
    ]

    # 1. Scan explicit classification columns first
    matched_explicit = []
    for col in raw_row.index:
        col_str = str(col).lower().strip().replace("_", " ").replace("-", " ")
        for h in explicit_headers:
            if h == col_str or h in col_str:
                matched_explicit.append(col)
                break

    for col in matched_explicit:
        val = str(raw_row[col]).strip().lower()
        if val:
            for kw in SORTED_KEYWORDS:
                if word_contains_kw(val, kw):
                    return KEYWORDS_MAP[kw]

    # 2. Fall back to description/reference columns only if explicit mapping didn't find anything
    matched_fallback = []
    for col in raw_row.index:
        col_str = str(col).lower().strip().replace("_", " ").replace("-", " ")
        for h in fallback_headers:
            if h == col_str or h in col_str:
                matched_fallback.append(col)
                break

    for col in matched_fallback:
        val = str(raw_row[col]).strip().lower()
        if val:
            for kw in SORTED_KEYWORDS:
                if word_contains_kw(val, kw):
                    return KEYWORDS_MAP[kw]

    # 3. Ultimate Fallback — Scan all remaining non-numeric columns in row
    for col in raw_row.index:
        val = str(raw_row[col]).strip().lower()
        if val and not _looks_numeric(val):
            for kw in SORTED_KEYWORDS:
                if word_contains_kw(val, kw):
                    return KEYWORDS_MAP[kw]

    return "Unknown"


EXCEL_EPOCH = pd.Timestamp("1899-12-30")


def _excel_serial_to_date(serial):
    try:
        serial = float(serial)
    except (TypeError, ValueError):
        return pd.NaT
    if serial < 18000 or serial > 73000:
        return pd.NaT
    try:
        return EXCEL_EPOCH + pd.to_timedelta(serial, unit="D")
    except (OverflowError, ValueError):
        return pd.NaT


def parse_date_series(series):
    """
    Vectorized parse dates prioritizing ISO layouts, UK formats, Excel serials.
    """
    s = series.astype(str).str.strip()
    result = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    empty_mask = series.isna() | (s == "") | (s.str.lower() == "nan")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)

        serial_mask = (~empty_mask) & s.str.fullmatch(r"\d{4,6}")
        if serial_mask.any():
            result.loc[serial_mask] = s[serial_mask].apply(_excel_serial_to_date)

        remaining = result.isna() & ~empty_mask
        if remaining.any():
            result.loc[remaining] = pd.to_datetime(s[remaining], errors="coerce", dayfirst=False)

        remaining = result.isna() & ~empty_mask
        if remaining.any():
            result.loc[remaining] = pd.to_datetime(s[remaining], errors="coerce", dayfirst=True)

        remaining = result.isna() & ~empty_mask
        if remaining.any():
            from dateutil import parser as _dtparser
            def _fuzzy(v):
                try:
                    return pd.Timestamp(_dtparser.parse(v, dayfirst=False, fuzzy=True))
                except (ValueError, OverflowError):
                    try:
                        return pd.Timestamp(_dtparser.parse(v, dayfirst=True, fuzzy=True))
                    except Exception:
                        return pd.NaT
                except Exception:
                    return pd.NaT

            result.loc[remaining] = s[remaining].apply(_fuzzy)

    return result


def format_date_display(val):
    if pd.isna(val):
        return ""
    return val.strftime("%Y-%m-%d")


def build_standard_rows(df, mapping):
    """
    Unifies disparate sheet inputs into standard matching rows.
    Also guesses date and description columns strictly for exact duplicate validation
    when unselected by the user. Supports dual amount column mapping (Amount and Amount 2).
    Saves original untranslated data to guarantee exact exception mapping.
    """
    rows = []
    amount_col = mapping.get("Amount")
    amount2_col = mapping.get("Amount 2")
    date_col = mapping.get("Date") if mapping.get("Date") != BLANK_OPTION else None
    ref_col = mapping.get("Reference") if mapping.get("Reference") != BLANK_OPTION else None
    desc_col = mapping.get("Description") if mapping.get("Description") != BLANK_OPTION else None

    # Guessing column candidates strictly for duplicate evaluation if unselected
    cols_list = list(df.columns.astype(str))
    guessed_date_col = date_col if date_col else auto_detect_column(cols_list, "Date")
    guessed_desc_col = desc_col if desc_col else auto_detect_column(cols_list, "Description")

    # Match raw type prioritized columns list (checking explicitly for type keywords first)
    matched_type_cols = []
    for col in df.columns:
        col_str = str(col).lower().strip()
        if any(kw in col_str for kw in ["type", "category", "class"]):
            matched_type_cols.append(col)

    if not matched_type_cols:
        priority_headers = [
            "description", "transection type", "transection_type", "transection types",
            "trasection_type", "trasection type", "transaction type", "transaction_type",
            "transaction types", "document type", "document_type", "reference", "ref"
        ]
        for col in df.columns:
            col_str = str(col).lower().strip().replace("_", " ").replace("-", " ")
            for p_kw in priority_headers:
                if p_kw == col_str or p_kw in col_str:
                    matched_type_cols.append(col)
                    break

    dates_series = parse_date_series(df[date_col]) if date_col else pd.Series(pd.NaT, index=df.index)

    # Generate guess-based parsed dates securely
    if guessed_date_col and guessed_date_col != date_col:
        guessed_dates_series = parse_date_series(df[guessed_date_col])
    else:
        guessed_dates_series = dates_series

    for idx, raw_row in df.iterrows():
        # Detect row category first to guide dual-column selection
        cat = detect_row_category(raw_row)

        # Resolve raw amount values based on dual column mapping and transaction category
        raw_amt = raw_row.get(amount_col) if amount_col else None
        raw_amt2 = raw_row.get(amount2_col) if amount2_col else None

        blank1 = is_blank_value(raw_amt)
        blank2 = is_blank_value(raw_amt2)

        final_raw_amt = None
        is_amt_blank = True
        is_dual_blank_payment = False

        if amount_col and amount2_col:
            # If both columns are blank/missing
            if blank1 and blank2:
                is_dual_blank_payment = True
                is_amt_blank = True
                final_raw_amt = None
            else:
                # Dual amount columns mapped: align by transaction type dynamically
                if cat == "Payment" and not blank2:
                    final_raw_amt = raw_amt2
                    is_amt_blank = False
                elif cat == "Invoice" and not blank1:
                    final_raw_amt = raw_amt
                    is_amt_blank = False
                else:
                    # Fallback to whichever contains a value
                    if not blank1:
                        final_raw_amt = raw_amt
                        is_amt_blank = False
                    elif not blank2:
                        final_raw_amt = raw_amt2
                        is_amt_blank = False
        elif amount_col:
            final_raw_amt = raw_amt
            is_amt_blank = blank1
        elif amount2_col:
            final_raw_amt = raw_amt2
            is_amt_blank = blank2

        amt = parse_amount(final_raw_amt) if final_raw_amt is not None else 0.0

        # CRITICAL FILTER: Ignore non-blank mapped values strictly less than 1.0 (such as 0.9, 0.8, etc.)
        # Blank amounts are kept so they can be flagged cleanly as "Amount not available" downstream.
        if not is_amt_blank and amt < 1.0:
            continue

        # SKIP LOGIC: If transaction type is Payment and its parsed amount is 0.0 or the column is empty, skip entirely!
        if cat == "Payment":
            if amt == 0.0:
                continue
            if amount2_col and blank2:
                continue
            elif not amount2_col and blank1:
                continue

        dt = dates_series.iloc[idx] if date_col else pd.NaT
        ref = str(raw_row.get(ref_col)).strip() if ref_col and pd.notna(raw_row.get(ref_col)) else ""
        desc = str(raw_row.get(desc_col)).strip() if desc_col and pd.notna(raw_row.get(desc_col)) else ""

        # Populate checked values for duplicate folder matches (guarantees exact criteria)
        chk_dt = guessed_dates_series.iloc[idx] if guessed_date_col else pd.NaT
        chk_dt_str = format_date_display(chk_dt) if pd.notna(chk_dt) else ""
        chk_desc = str(raw_row.get(guessed_desc_col)).strip() if guessed_desc_col and pd.notna(
            raw_row.get(guessed_desc_col)) else ""

        # Obtain 1-based physical raw Excel/CSV row number
        original_row_num = int(raw_row.get("original_row_num", idx + 1))

        # Save exact unmodified physical cell values for accurate display
        orig_amount_str = str(final_raw_amt) if final_raw_amt is not None else ""
        if amount_col and amount2_col and is_dual_blank_payment:
            orig_amount_str = ""

        orig_date_str = str(raw_row.get(date_col)) if date_col and pd.notna(raw_row.get(date_col)) else ""
        orig_ref_str = str(raw_row.get(ref_col)) if ref_col and pd.notna(raw_row.get(ref_col)) else ""
        orig_desc_str = str(raw_row.get(desc_col)) if desc_col and pd.notna(raw_row.get(desc_col)) else ""

        # Extract the raw, original transaction type string from mapped prioritized headers
        raw_type_val = ""
        for col in matched_type_cols:
            val_str = str(raw_row.get(col, "")).strip()
            if val_str:
                raw_type_val = val_str
                break

        rows.append({
            "idx": idx,
            "original_row_num": original_row_num,
            "amount": amt,
            "date": dt,
            "reference": ref,
            "description": desc,
            "category": cat,
            "is_blank_amount": is_amt_blank,
            "is_dual_blank_payment": is_dual_blank_payment,
            "check_date": chk_dt_str,
            "check_description": chk_desc,
            # Original raw sheet metadata for consistent error reporting
            "orig_amount": orig_amount_str,
            "orig_date": orig_date_str,
            "orig_ref": orig_ref_str,
            "orig_desc": orig_desc_str,
            "raw_type": raw_type_val if raw_type_val else "Transection type is missing"
        })
    return rows


# ==========================================================================
# BLOCK 6: WEIGHTED MATCHING SCORING ENGINE (90% AMOUNT, 5% DATE, 5% REF)
# ==========================================================================
def score_pair(a, b, show_date, show_ref):
    """
    Score validation engine (Max 100):
    - Exact Amount match (diff <= 0.02) = 90 pts
    - Same Date = 5 pts, within 3 days = 3 pts, within 10 days = 1 pt (only evaluated if Date mapped in both)
    - Exact Reference match = 5 pts (only evaluated if Reference mapped in both)
    """
    # Guard clause: do not match if either has amount missing
    if a.get("is_blank_amount", False) or b.get("is_blank_amount", False):
        return 0, []

    amt_diff = abs(a["amount"] - b["amount"])
    if amt_diff > 0.02:
        return 0, []

    score = 90
    basis = ["Amount exact match (+90)"]

    if show_date and pd.notna(a["date"]) and pd.notna(b["date"]):
        days = abs((a["date"] - b["date"]).days)
        if days == 0:
            score += 5
            basis.append("Same-day date (+5)")
        elif days <= 3:
            score += 3
            basis.append(f"Date within {days} days (+3)")
        elif days <= 10:
            score += 1
            basis.append(f"Date within {days} days (+1)")
        else:
            basis.append(f"Date mismatch ({days} days) (+0)")

    if show_ref:
        ra = str(a["reference"]).strip().lower()
        rb = str(b["reference"]).strip().lower()
        if ra and rb and ra != "nan" and rb != "nan" and ra == rb:
            score += 5
            basis.append("Reference exact match (+5)")

    return score, basis


# ==========================================================================
# BLOCK 7: MULTI-PASS RECONCILIATION ENGINE
# ==========================================================================
def run_reconciliation(df1_orig, df2_orig, mapping1, mapping2, name1, name2):
    """
    Performs comprehensive intra-sheet offset mapping, cross-sheet direct matching,
    combinatorial sum grouping (1-to-Many, Many-to-1), and error classification.
    """
    show_date = mapping1.get("Date") is not None and mapping2.get("Date") is not None
    show_ref = mapping1.get("Reference") is not None and mapping2.get("Reference") is not None

    rows1 = build_standard_rows(df1_orig, mapping1)
    rows2 = build_standard_rows(df2_orig, mapping2)

    categories = ["Invoice", "Payment", "Credit Note", "Adjustment", "Unknown"]
    totals1 = {cat: 0.0 for cat in categories}
    totals2 = {cat: 0.0 for cat in categories}
    counts1 = {cat: 0 for cat in categories}
    counts2 = {cat: 0 for cat in categories}

    for r in rows1:
        if not r.get("is_blank_amount", False):
            totals1[r["category"]] += r["amount"]
        counts1[r["category"]] += 1

    for r in rows2:
        if not r.get("is_blank_amount", False):
            totals2[r["category"]] += r["amount"]
        counts2[r["category"]] += 1

    # Standard Accounting Balance Policy matching Title Case indicators exactly:
    # Starting from absolute 0. Payment, Credit Note, Adjustment add (+). Invoice subtract (-). Unknown excluded.
    bal1 = 0.0 + totals1["Payment"] + totals1["Credit Note"] + totals1["Adjustment"] - totals1["Invoice"]
    bal2 = 0.0 + totals2["Payment"] + totals2["Credit Note"] + totals2["Adjustment"] - totals2["Invoice"]
    balance_diff = bal1 - bal2

    # Extract all unclassified Unknown records for distinct audit folder isolation
    unclassified1 = [r for r in rows1 if r["category"] == "Unknown"]
    unclassified2 = [r for r in rows2 if r["category"] == "Unknown"]

    intra_sheet_matches1 = []
    intra_sheet_matches2 = []
    intra_used1 = set()
    intra_used2 = set()

    # ---- PASS 0: Internal self-offsets inside Sheet 1 ----
    for i in range(len(rows1)):
        if i in intra_used1 or rows1[i].get("is_blank_amount", False):
            continue
        r1_a = rows1[i]
        for j in range(i + 1, len(rows1)):
            if j in intra_used1 or rows1[j].get("is_blank_amount", False):
                continue
            r1_b = rows1[j]

            if abs(r1_a["amount"] - r1_b["amount"]) <= 0.01:
                is_offset = False
                if (r1_a["category"] == "Payment" and r1_b["category"] in ("Credit Note", "Adjustment")) or \
                        (r1_b["category"] == "Payment" and r1_a["category"] in ("Credit Note", "Adjustment")):
                    is_offset = True
                elif (r1_a["category"] == "Invoice" and r1_b["category"] in ("Credit Note", "Adjustment")) or \
                        (r1_b["category"] == "Invoice" and r1_a["category"] in ("Credit Note", "Adjustment")):
                    is_offset = True

                if is_offset:
                    intra_used1.update([i, j])
                    intra_sheet_matches1.append({
                        "row_a": r1_a,
                        "row_b": r1_b,
                        "amount": r1_a["amount"],
                        "type": "Intra-Sheet Offset"
                    })
                    break

    # ---- PASS 0: Internal self-offsets inside Sheet 2 ----
    for i in range(len(rows2)):
        if i in intra_used2 or rows2[i].get("is_blank_amount", False):
            continue
        r2_a = rows2[i]
        for j in range(i + 1, len(rows2)):
            if j in intra_used2 or rows2[j].get("is_blank_amount", False):
                continue
            r2_b = rows2[j]

            if abs(r2_a["amount"] - r2_b["amount"]) <= 0.01:
                is_offset = False
                if (r2_a["category"] == "Payment" and r2_b["category"] in ("Credit Note", "Adjustment")) or \
                        (r2_b["category"] == "Payment" and r2_a["category"] in ("Credit Note", "Adjustment")):
                    is_offset = True
                elif (r2_a["category"] == "Invoice" and r2_b["category"] in ("Credit Note", "Adjustment")) or \
                        (r2_b["category"] == "Invoice" and r2_a["category"] in ("Credit Note", "Adjustment")):
                    is_offset = True

                if is_offset:
                    intra_used2.update([i, j])
                    intra_sheet_matches2.append({
                        "row_a": r2_a,
                        "row_b": r2_b,
                        "amount": r2_a["amount"],
                        "type": "Intra-Sheet Offset"
                    })
                    break

    inter_sheet_matches = []

    # ---- STEP 1: Match Same Category transactions first (Invoice vs Invoice, Payment vs Payment, etc.) ----
    same_type_matches = []
    matched_same1 = set()
    matched_same2 = set()

    for i, r1 in enumerate(rows1):
        if i in intra_used1 or r1.get("is_blank_amount", False):
            continue
        for j, r2 in enumerate(rows2):
            if j in intra_used2 or r2.get("is_blank_amount", False):
                continue

            if r1["category"] == r2["category"] and r1["category"] != "Unknown":
                score, basis = score_pair(r1, r2, show_date, show_ref)
                if score >= 90:
                    same_type_matches.append({
                        "score": score, "i": i, "j": j,
                        "type": "Direct Match", "basis": basis
                    })

    same_type_matches = sorted(same_type_matches, key=lambda x: -x["score"])
    for cand in same_type_matches:
        if cand["i"] in matched_same1 or cand["j"] in matched_same2:
            continue
        matched_same1.add(cand["i"])
        matched_same2.add(cand["j"])
        inter_sheet_matches.append({
            "row1": rows1[cand["i"]], "row2": rows2[cand["j"]],
            "score": cand["score"], "type": "Direct Match", "basis": "; ".join(cand["basis"])
        })

    # ---- STEP 2: Match Cross Offset Categories without exclusion ----
    cross_type_matches = []
    matched_cross1 = set()
    matched_cross2 = set()

    for i, r1 in enumerate(rows1):
        if i in intra_used1 or r1.get("is_blank_amount", False):
            continue
        for j, r2 in enumerate(rows2):
            if j in intra_used2 or r2.get("is_blank_amount", False):
                continue

            # Multi-category offset matching rules: match any different transaction category groupings
            valid_categories = {"Invoice", "Payment", "Credit Note", "Adjustment"}
            if r1["category"] in valid_categories and r2["category"] in valid_categories:
                if r1["category"] != r2["category"]:
                    score, basis = score_pair(r1, r2, show_date, show_ref)
                    if score >= 90:
                        cross_type_matches.append({
                            "score": score, "i": i, "j": j,
                            "type": "Cross-Offset Match", "basis": basis
                        })

    cross_type_matches = sorted(cross_type_matches, key=lambda x: -x["score"])
    for cand in cross_type_matches:
        if cand["i"] in matched_cross1 or cand["j"] in matched_cross2:
            continue
        matched_cross1.add(cand["i"])
        matched_cross2.add(cand["j"])
        inter_sheet_matches.append({
            "row1": rows1[cand["i"]], "row2": rows2[cand["j"]],
            "score": cand["score"], "type": "Cross-Offset Match", "basis": "; ".join(cand["basis"])
        })

    # ---- PASS 3: One-to-Many Combinatorial Sum Checks ----
    leftover1 = [i for i in range(len(rows1)) if
                 i not in intra_used1 and i not in matched_same1 and i not in matched_cross1]
    leftover2 = [j for j in range(len(rows2)) if
                 j not in intra_used2 and j not in matched_same2 and j not in matched_cross2]
    advanced_matches = []

    for i in leftover1:
        r1 = rows1[i]
        if r1.get("is_blank_amount", False):
            continue
        pool = [j for j in leftover2 if not rows2[j].get("is_blank_amount", False)][:30]

        found = None
        for r in range(2, 4):
            for combo in itertools.combinations(pool, r):
                combo_sum = sum(rows2[j]["amount"] for j in combo)
                if abs(r1["amount"] - combo_sum) <= 0.05:
                    found = combo
                    break
            if found:
                break

        if found:
            for j in found:
                if j in leftover2:
                    leftover2.remove(j)
            advanced_matches.append({
                "row1": r1, "rows2": [rows2[j] for j in found],
                "score": 90,
                "type": "One-to-Many Sum Match",
                "basis": "Matched sum of items"
            })

    # ---- PASS 3: Many-to-One Combinatorial Sum Checks ----
    leftover1 = [i for i in range(len(rows1)) if
                 i not in intra_used1 and i not in matched_same1 and i not in matched_cross1]
    leftover2 = [j for j in range(len(rows2)) if
                 j not in intra_used2 and j not in matched_same2 and j not in matched_cross2]

    for j in leftover2:
        r2 = rows2[j]
        if r2.get("is_blank_amount", False):
            continue
        pool = [i for i in leftover1 if not rows1[i].get("is_blank_amount", False)][:30]

        found = None
        for r in range(2, 4):
            for combo in itertools.combinations(pool, r):
                combo_sum = sum(rows1[i]["amount"] for i in combo)
                if abs(r2["amount"] - combo_sum) <= 0.05:
                    found = combo
                    break
            if found:
                break

        if found:
            for i in found:
                if i in leftover1:
                    leftover1.remove(i)
            advanced_matches.append({
                "row2": r2, "rows1": [rows1[i] for i in found],
                "score": 90,
                "type": "Many-to-One Sum Match",
                "basis": "Matched sum of items"
            })

    unmatched1 = [i for i in range(len(rows1)) if
                  i not in intra_used1 and i not in matched_same1 and i not in matched_cross1]
    unmatched2 = [j for j in range(len(rows2)) if
                  j not in intra_used2 and j not in matched_same2 and j not in matched_cross2]

    def separate_exceptions_and_duplicates(rows_list, unmatched_idxs, opposite_file_name):
        """
        Groups unmatched rows. Flags a transaction as a duplicate ONLY when the
        Amount, Date, and Description fields match exactly.
        Applies specific exception reasons referring to the opposite missing file.
        """
        freq = {}
        for idx in unmatched_idxs:
            r = rows_list[idx]
            if r.get("is_blank_amount", False):
                continue

            key = (r["amount"], r["check_date"], r["check_description"].lower().strip())
            freq[key] = freq.get(key, 0) + 1

        exceptions = []
        duplicates = []
        for idx in unmatched_idxs:
            r = rows_list[idx]
            r_copy = r.copy()
            if r.get("is_blank_amount", False):
                if r.get("is_dual_blank_payment", False):
                    r_copy["exception_type"] = "Missing Payment Category"
                else:
                    r_copy["exception_type"] = "Amount not available"
                exceptions.append(r_copy)
            else:
                key = (r["amount"], r["check_date"], r["check_description"].lower().strip())
                if freq[key] > 1:
                    r_copy["exception_type"] = f"Duplicate {r['category']}"
                    duplicates.append(r_copy)
                else:
                    # Provide exact contextual reason missing in opposite file
                    cat_val = r['category']
                    if cat_val == "Unknown":
                        cat_val = "Unclassified Transaction"
                    r_copy["exception_type"] = f"{cat_val} missing in {opposite_file_name}"
                    exceptions.append(r_copy)
        return exceptions, duplicates

    exceptions1, dupes1 = separate_exceptions_and_duplicates(rows1, unmatched1, name2)
    exceptions2, dupes2 = separate_exceptions_and_duplicates(rows2, unmatched2, name1)

    return {
        "totals1": totals1, "totals2": totals2,
        "counts1": counts1, "counts2": counts2,
        "bal1": bal1, "bal2": bal2, "bal_diff": balance_diff,
        "unclassified1": unclassified1, "unclassified2": unclassified2,
        "intra1": intra_sheet_matches1, "intra2": intra_sheet_matches2,
        "inter_matches": inter_sheet_matches,
        "advanced_matches": advanced_matches,
        "exceptions1": exceptions1, "exceptions2": exceptions2,
        "dupes1": dupes1, "dupes2": dupes2,
        "len_rows1": len(rows1),
        "len_rows2": len(rows2)
    }


# ==========================================================================
# BLOCK 8: EXCEL FILE EXPORT ENGINE
# ==========================================================================
def write_cell_value(cell, val):
    """
    Writes values as formatted numbers if they look like digits or currency blocks,
    preserving plain text format otherwise. Resolves parenthesized formats cleanly.
    Does not write additional trailing decimal zeros (.00) on clean integers/whole numbers.
    """
    if pd.isna(val):
        cell.value = ""
        return

    val_str = str(val).strip()

    # If already native float or int, write directly
    if isinstance(val, (int, float)):
        cell.value = val
        if isinstance(val, int) or (isinstance(val, float) and val.is_integer()):
            cell.number_format = '#,##0'
        else:
            cell.number_format = '#,##0.00'
        return

    # Clear comma and currency indicators
    clean_str = val_str.replace("$", "").replace(",", "")
    is_neg = False
    if clean_str.startswith("(") and clean_str.endswith(")"):
        clean_str = clean_str[1:-1]
        is_neg = True
    elif clean_str.startswith("-"):
        clean_str = clean_str[1:]
        is_neg = True

    # Check if this string qualifies as pure float/integer digits
    if clean_str.replace(".", "", 1).isdigit() and clean_str.count(".") <= 1:
        try:
            num_val = float(clean_str)
            if is_neg:
                num_val = -num_val

            if num_val.is_integer():
                cell.value = int(num_val)
                if "$" in val_str:
                    cell.number_format = '$#,##0'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.value = num_val
                if "$" in val_str:
                    cell.number_format = '$#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
            return
        except ValueError:
            pass

    cell.value = val_str


def generate_excel_report(tables_list):
    """
    Generates a single, beautiful sequential Excel workbook.
    Headers are written as non-colored bold black text.
    Numbers are outputted as actual numeric cell values.
    Skips any empty summary or matching tables.
    """
    import io
    output = io.BytesIO()
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reconciliation Report"

        # High contrast, plain black/bold formatting for headings (No Colors)
        title_font = Font(name='Segoe UI', size=14, bold=True, color='000000')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='000000')

        current_row = 1
        for title, df in tables_list:
            if df is None or df.empty:
                continue

            # Write Title header as bold black text
            ws.cell(row=current_row, column=1, value=title).font = title_font
            current_row += 2

            # Write column headers as bold black text without color background fills
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Write dataframe cell values (storing numbers natively)
            for _, row in df.iterrows():
                for c_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=current_row, column=c_idx)
                    write_cell_value(cell, val)
                current_row += 1

            # Add spacer pad rows
            current_row += 3

        wb.save(output)
    except Exception:
        # Fallback exporter utilizing basic xlsxwriter
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            current_row = 0
            for title, df in tables_list:
                if df is None or df.empty:
                    continue
                df.to_excel(writer, sheet_name="Reconciliation Report", startrow=current_row + 1, index=False)
                current_row += len(df) + 4

    return output.getvalue()


def build_excel_tables_tracker(R, name1, name2):
    """
    Sequentially compiles standard reporting dataframes for Excel download.
    Strictly excludes the unclassified transactions report as requested.
    """
    tracker = []
    show_date = R["mapping1"].get("Date") is not None and R["mapping2"].get("Date") is not None
    show_ref = R["mapping1"].get("Reference") is not None and R["mapping2"].get("Reference") is not None
    show_desc = R["mapping1"].get("Description") is not None and R["mapping2"].get("Description") is not None

    # 1. Data Summary
    categories_list = ["Invoice", "Payment", "Credit Note", "Adjustment", "Unknown"]
    summary_data = []
    for cat in categories_list:
        total1 = R["totals1"][cat]
        cnt1 = R["counts1"][cat]
        total2 = R["totals2"][cat]
        cnt2 = R["counts2"][cat]
        diff_val = total1 - total2
        clean_cat_lbl = "Unclassified" if cat == "Unknown" else cat
        summary_data.append({
            "Transaction Category": clean_cat_lbl,
            f"{name1} Count": cnt1,
            f"{name1} Total Amount": f"${total1:,.2f}",
            f"{name2} Count": cnt2,
            f"{name2} Total Amount": f"${total2:,.2f}",
            "Count Difference": abs(cnt1 - cnt2),
            "Absolute Difference": f"${abs(diff_val):,.2f}",
            "Status Variance": "Perfect Align" if abs(diff_val) <= 0.01 else (
                f"${abs(diff_val):,.2f} missing on {name2}" if diff_val > 0 else f"${abs(diff_val):,.2f} missing on {name1}")
        })
    tracker.append(("Data Summary", pd.DataFrame(summary_data)))

    # 2. Reconciliation Results
    n_direct = len(R["inter_matches"])
    one_to_many_matches = [m for m in R["advanced_matches"] if "row1" in m]
    many_to_one_matches = [m for m in R["advanced_matches"] if "row2" in m]
    n_one_to_many = len(one_to_many_matches)
    n_many_to_one = len(many_to_one_matches)
    n_intra = len(R["intra1"]) + len(R["intra2"])
    n_exceptions = len(R["exceptions1"]) + len(R["exceptions2"])
    n_dupes = len(R["dupes1"]) + len(R["dupes2"])

    matched_indices1, matched_indices2 = set(), set()
    for m in R["inter_matches"]:
        matched_indices1.add(m["row1"]["idx"])
        matched_indices2.add(m["row2"]["idx"])
    for m in R["advanced_matches"]:
        if "row1" in m:
            matched_indices1.add(m["row1"]["idx"])
            for r in m["rows2"]:
                matched_indices2.add(r["idx"])
        else:
            matched_indices2.add(m["row2"]["idx"])
            for r in m["rows1"]:
                matched_indices1.add(r["idx"])
    for m in R["intra1"]:
        matched_indices1.add(m["row_a"]["idx"])
        matched_indices1.add(m["row_b"]["idx"])
    for m in R["intra2"]:
        matched_indices2.add(m["row_a"]["idx"])
        matched_indices2.add(m["row_b"]["idx"])

    total_unique_matched = len(matched_indices1) + len(matched_indices2)
    total_analyzed = R["len_rows1"] + R["len_rows2"]
    percentage_matched = (total_unique_matched / total_analyzed * 100) if total_analyzed > 0 else 0.0

    results_df = pd.DataFrame([
        {"Metric": "Total Records Analyzed in both files", "Result": str(total_analyzed)},
        {"Metric": "Successfully Matched", "Result": str(n_direct)},
        {"Metric": "% of Successfully Matched Transactions", "Result": f"{percentage_matched:.2f}%"},
        {"Metric": "One-to-Many Sum Matched", "Result": str(n_one_to_many)},
        {"Metric": "Many-to-One Sum Matched", "Result": str(n_many_to_one)},
        {"Metric": "Intra-Sheet Offsets", "Result": str(n_intra)},
        {"Metric": "Exceptions Unmatched", "Result": str(n_exceptions)},
        {"Metric": "Possible Duplicates", "Result": str(n_dupes)}
    ])
    tracker.append(("Reconciliation Results", results_df))

    # 3. Genuine Cross-Sheet Matched Transactions
    if n_direct:
        matched_rows = []
        for m in R["inter_matches"]:
            r1, r2 = m["row1"], m["row2"]
            matched_props = ["Amount"]
            unmatched_props = []

            if show_date:
                if pd.notna(r1["date"]) and pd.notna(r2["date"]):
                    days = abs((r1["date"] - r2["date"]).days)
                    matched_props.append("Date") if days == 0 else unmatched_props.append(f"Date ({days} days diff)")
                else:
                    unmatched_props.append("Date blank in record")

            if show_ref:
                ref1, ref2 = clean_cell_text(r1["reference"]), clean_cell_text(r2["reference"])
                if ref1 and ref2:
                    matched_props.append("Reference") if ref1 == ref2 else unmatched_props.append("Reference mismatch")
                else:
                    unmatched_props.append("Reference blank in record")

            status_str = "Matched: " + ", ".join(matched_props)
            if unmatched_props:
                status_str += " | Unmatched: " + ", ".join(unmatched_props)

            row_dict = {
                f"{name1}: Row": r1["original_row_num"],
                f"{name2}: Row": r2["original_row_num"],
                f"{name1}: Transaction Type": "" if r1["category"] == "Unknown" else r1["category"],
                f"{name2}: Transaction Type": "" if r2["category"] == "Unknown" else r2["category"],
                f"{name1}: Amount": clean_amount_display(r1["amount"]),
                f"{name2}: Amount": clean_amount_display(r2["amount"]),
            }
            if show_date:
                row_dict[f"{name1}: Date"] = format_date_display(r1["date"])
                row_dict[f"{name2}: Date"] = format_date_display(r2["date"])
            if show_ref:
                row_dict[f"{name1}: Reference"] = ref1
                row_dict[f"{name2}: Reference"] = ref2
            if show_desc:
                row_dict[f"{name1}: Description"] = clean_cell_text(r1["description"])
                row_dict[f"{name2}: Description"] = clean_cell_text(r2["description"])

            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = status_str
            matched_rows.append(row_dict)
        tracker.append(("Genuine Cross-Sheet Matched Transactions", pd.DataFrame(matched_rows)))

    # 4. Intra-Sheet Self-Offsets
    if n_intra:
        intra_rows = []
        for m in R["intra1"]:
            cat_a = m['row_a']['category'] if m['row_a']['category'] != 'Unknown' else 'Unclassified'
            cat_b = m['row_b']['category'] if m['row_b']['category'] != 'Unknown' else 'Unclassified'
            row_dict = {
                "Statement File": name1,
                "Offset Rows": f"Row {m['row_a']['original_row_num']} & Row {m['row_b']['original_row_num']}",
                "Amount": clean_amount_display(m["amount"]),
                "Reconciled Category Types": f"{cat_a} vs {cat_b}",
                "Reconciliation Basis": m["type"],
                "Reconciliation Status": "Matched Internally within File"
            }
            if show_date:
                row_dict[
                    "Dates"] = f"{format_date_display(m['row_a']['date'])}, {format_date_display(m['row_b']['date'])}".strip(
                    ", —")
            if show_ref:
                row_dict[
                    "References"] = f"{clean_cell_text(m['row_a']['reference'])}, {clean_cell_text(m['row_b']['reference'])}".strip(
                    ", ")
            intra_rows.append(row_dict)

        for m in R["intra2"]:
            cat_a = m['row_a']['category'] if m['row_a']['category'] != 'Unknown' else 'Unclassified'
            cat_b = m['row_b']['category'] if m['row_b']['category'] != 'Unknown' else 'Unclassified'
            row_dict = {
                "Statement File": name2,
                "Offset Rows": f"Row {m['row_a']['original_row_num']} & Row {m['row_b']['original_row_num']}",
                "Amount": clean_amount_display(m["amount"]),
                "Reconciled Category Types": f"{cat_a} vs {cat_b}",
                "Reconciliation Basis": m["type"],
                "Reconciliation Status": "Matched Internally within File"
            }
            if show_date:
                row_dict[
                    "Dates"] = f"{format_date_display(m['row_a']['date'])}, {format_date_display(m['row_b']['date'])}".strip(
                    ", —")
            if show_ref:
                row_dict[
                    "References"] = f"{clean_cell_text(m['row_a']['reference'])}, {clean_cell_text(m['row_b']['reference'])}".strip(
                    ", ")
            intra_rows.append(row_dict)
        tracker.append(("Intra-Sheet Self-Offsets (Internal Adjustments)", pd.DataFrame(intra_rows)))

    # 5. Split Payments (One-to-Many)
    if n_one_to_many:
        otm_rows = []
        for m in one_to_many_matches:
            r1, rows2 = m["row1"], m["rows2"]
            r2_idxs = ", ".join(str(r["original_row_num"]) for r in rows2)
            r2_types = ", ".join(r["category"] for r in rows2 if r["category"] != "Unknown")
            r2_refs = ", ".join(clean_cell_text(r["reference"]) for r in rows2 if clean_cell_text(r["reference"]))
            r2_amts = sum(r["amount"] for r in rows2)
            r2_dates = ", ".join(format_date_display(r["date"]) for r in rows2 if pd.notna(r["date"]))
            r2_descs = ", ".join(clean_cell_text(r["description"]) for r in rows2 if clean_cell_text(r["description"]))
            row_dict = {
                f"{name1}: Row": r1["original_row_num"],
                f"{name2}: Row(s)": r2_idxs,
                f"{name1}: Transaction Type": "" if r1["category"] == "Unknown" else r1["category"],
                f"{name2}: Transaction Type(s)": r2_types,
                f"{name1}: Amount": clean_amount_display(r1["amount"]),
                f"{name2}: Amount (Sum)": clean_amount_display(r2_amts),
            }
            if show_date:
                row_dict[f"{name1}: Date"] = format_date_display(r1["date"])
                row_dict[f"{name2}: Date(s)"] = r2_dates
            if show_ref:
                row_dict[f"{name1}: Reference"] = clean_cell_text(r1["reference"])
                row_dict[f"{name2}: Reference(s)"] = r2_refs
            if show_desc:
                row_dict[f"{name1}: Description"] = clean_cell_text(r1["description"])
                row_dict[f"{name2}: Description(s)"] = r2_descs
            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = "Matched: Amount (Sum) | Individual references or dates differ across items"
            otm_rows.append(row_dict)
        tracker.append(("Split Payments (One-to-Many Ledger Reconciliation)", pd.DataFrame(otm_rows)))

    # 6. Combined Payments (Many-to-One)
    if n_many_to_one:
        mto_rows = []
        for m in many_to_one_matches:
            r2, rows1 = m["row2"], m["rows1"]
            r1_idxs = ", ".join(str(r["original_row_num"]) for r in rows1)
            r1_types = ", ".join(r["category"] for r in rows1 if r["category"] != "Unknown")
            r1_refs = ", ".join(clean_cell_text(r["reference"]) for r in rows1 if clean_cell_text(r["reference"]))
            r1_amts = sum(r["amount"] for r in rows1)
            r1_dates = ", ".join(format_date_display(r["date"]) for r in rows1 if pd.notna(r["date"]))
            r1_descs = ", ".join(clean_cell_text(r["description"]) for r in rows1 if clean_cell_text(r["description"]))
            row_dict = {
                f"{name1}: Row(s)": r1_idxs,
                f"{name2}: Row": r2["original_row_num"],
                f"{name1}: Transaction Type(s)": r1_types,
                f"{name2}: Transaction Type": "" if r2["category"] == "Unknown" else r2["category"],
                f"{name1}: Amount (Sum)": clean_amount_display(r1_amts),
                f"{name2}: Amount": clean_amount_display(r2["amount"]),
            }
            if show_date:
                row_dict[f"{name1}: Date(s)"] = r1_dates
                row_dict[f"{name2}: Date"] = format_date_display(r2["date"])
            if show_ref:
                row_dict[f"{name1}: Reference(s)"] = r1_refs
                row_dict[f"{name2}: Reference"] = clean_cell_text(r2["reference"])
            if show_desc:
                row_dict[f"{name1}: Description(s)"] = r1_descs
                row_dict[f"{name2}: Description"] = clean_cell_text(r2["description"])
            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = "Matched: Amount (Sum) | Individual references or dates differ across items"
            mto_rows.append(row_dict)
        tracker.append(("Combined Payments (Many-to-One Ledger Reconciliation)", pd.DataFrame(mto_rows)))

    # 7. Potential Duplicate Entries
    if n_dupes:
        dup_rows = []
        for dup in R["dupes1"]:
            dup_rows.append({
                "Source Statement": name1, "Row": dup["original_row_num"],
                "Amount": clean_cell_text(dup["orig_amount"]) if dup["orig_amount"] else "—",
                "Date": dup["orig_date"] if dup["orig_date"] else "—",
                "Description": dup["orig_desc"] if dup["orig_desc"] else "—",
                "Reference": clean_cell_text(dup["orig_ref"]) if show_ref else "—",
                "Category Type": "" if dup["category"] == "Unknown" else dup["category"],
                "Duplicate Category": dup["exception_type"],
                "_sort_amount": dup["amount"], "_sort_date": dup["check_date"],
                "_sort_desc": dup["check_description"].lower().strip()
            })
        for dup in R["dupes2"]:
            dup_rows.append({
                "Source Statement": name2, "Row": dup["original_row_num"],
                "Amount": clean_cell_text(dup["orig_amount"]) if dup["orig_amount"] else "—",
                "Date": dup["orig_date"] if dup["orig_date"] else "—",
                "Description": dup["orig_desc"] if dup["orig_desc"] else "—",
                "Reference": clean_cell_text(dup["orig_ref"]) if show_ref else "—",
                "Category Type": "" if dup["category"] == "Unknown" else dup["category"],
                "Duplicate Category": dup["exception_type"],
                "_sort_amount": dup["amount"], "_sort_date": dup["check_date"],
                "_sort_desc": dup["check_description"].lower().strip()
            })
        dup_df = pd.DataFrame(dup_rows)
        dup_df = dup_df.sort_values(by=["_sort_amount", "_sort_date", "_sort_desc", "Source Statement"],
                                    ascending=[True, True, True, True])
        dup_df = dup_df.drop(columns=["_sort_amount", "_sort_date", "_sort_desc"])
        tracker.append(("Potential Duplicate Entries", dup_df))

    # 7.5 EXCEPTION SUMMARY (Added to Excel Trackers)
    if n_exceptions:
        exc_summary = {
            "Missing Invoices": 0, "Missing Payments": 0, "Missing Credit Notes": 0, "Missing Adjustments": 0
        }
        for exc in R["exceptions1"] + R["exceptions2"]:
            cat = exc["category"]
            if cat == "Invoice":
                exc_summary["Missing Invoices"] += 1
            elif cat == "Payment":
                exc_summary["Missing Payments"] += 1
            elif cat == "Credit Note":
                exc_summary["Missing Credit Notes"] += 1
            elif cat == "Adjustment":
                exc_summary["Missing Adjustments"] += 1

        exc_summary_df = pd.DataFrame([
            {"Issue": "Missing Invoices", "Records": exc_summary["Missing Invoices"]},
            {"Issue": "Missing Payments", "Records": exc_summary["Missing Payments"]},
            {"Issue": "Missing Credit Notes", "Records": exc_summary["Missing Credit Notes"]},
            {"Issue": "Missing Adjustments", "Records": exc_summary["Missing Adjustments"]}
        ])
        tracker.append(("Exception Summary", exc_summary_df))

    # 8. Exceptions & Unmatched Report
    if n_exceptions:
        unmatched_rows = []
        for exc in R["exceptions1"]:
            row_dict = {
                "Source Statement": name1, "Row": exc["original_row_num"],
                "Amount": clean_cell_text(exc["orig_amount"]) if exc["orig_amount"] else "—",
            }
            if show_date:
                row_dict["Date"] = clean_cell_text(exc["orig_date"]) if exc["orig_date"] else "—"
            if show_ref:
                row_dict["Reference"] = clean_cell_text(exc["orig_ref"]) if exc["orig_ref"] else "—"
            row_dict.update({
                "Category Type": "" if exc["category"] == "Unknown" else exc["category"],
                "Description Context": clean_cell_text(exc["orig_desc"]) if exc["orig_desc"] else "—",
                "Exception Category": exc["exception_type"]
            })
            unmatched_rows.append(row_dict)
        for exc in R["exceptions2"]:
            row_dict = {
                "Source Statement": name2, "Row": exc["original_row_num"],
                "Amount": clean_cell_text(exc["orig_amount"]) if exc["orig_amount"] else "—",
            }
            if show_date:
                row_dict["Date"] = clean_cell_text(exc["orig_date"]) if exc["orig_date"] else "—"
            if show_ref:
                row_dict["Reference"] = clean_cell_text(exc["orig_ref"]) if exc["orig_ref"] else "—"
            row_dict.update({
                "Category Type": "" if exc["category"] == "Unknown" else exc["category"],
                "Description Context": clean_cell_text(exc["orig_desc"]) if exc["orig_desc"] else "—",
                "Exception Category": exc["exception_type"]
            })
            unmatched_rows.append(row_dict)
        exceptions_df = pd.DataFrame(unmatched_rows)
        exceptions_df["_sort_amount"] = exceptions_df["Amount"].apply(lambda x: parse_amount(x))
        exceptions_df = exceptions_df.sort_values(by=["_sort_amount", "Source Statement"], ascending=[True, True])
        exceptions_df = exceptions_df.drop(columns=["_sort_amount"])
        tracker.append(("Exceptions & Unmatched Report", exceptions_df))

    return tracker


# ==========================================================================
# BLOCK 9: STREAMLIT RENDERING HELPERS & FILE SLOTS
# ==========================================================================
def render_upload_slot(col, slot: int):
    name_key = f"file{slot}_name"
    bytes_key = f"file{slot}_bytes"
    uploader_key = f"uploader{slot}_key"

    with col:
        st.markdown(f'<div class="upload-label">File {slot}</div>', unsafe_allow_html=True)

        if st.session_state[name_key] is None:
            uploaded = st.file_uploader(
                "Upload file",
                type=list(ALLOWED_EXTENSIONS),
                key=f"uploader{slot}_{st.session_state[uploader_key]}",
                help="Accepted formats: CSV, XLSX, XLS, PDF",
                label_visibility="collapsed",
            )
            if uploaded is not None:
                ext = get_extension(uploaded.name)
                if ext in ALLOWED_EXTENSIONS:
                    st.session_state[name_key] = uploaded.name
                    st.session_state[bytes_key] = uploaded.getvalue()
                    st.rerun()
                else:
                    st.markdown(
                        f'<div class="file-err">❌ \'.{ext}\' is not supported. Please upload CSV, XLSX, XLS, or PDF.</div>',
                        unsafe_allow_html=True,
                    )
        else:
            ext = get_extension(st.session_state[name_key]).upper()
            info_col, del_col = st.columns([5, 1])
            with info_col:
                st.markdown(
                    f'<div class="file-ok">✅ <b>{st.session_state[name_key]}</b><span class="file-ext-badge">{ext}</span></div>',
                    unsafe_allow_html=True,
                )
            with del_col:
                if st.button("🗑️", key=f"delete_btn_{slot}", help="Remove this file"):
                    delete_file(slot)
                    st.rerun()


# ==========================================================================
# BLOCK 10: STREAMLIT UI ROUTING, MAPPING TABLE & DETAILED REPORTS
# ==========================================================================
st.markdown(
    """
    <div class="app-header">
        <h1>🗂️ Financial Reconciliation Tool</h1>
        <p>Smart transaction categorization, comparative balance calculations & prioritization-based matching</p>
    </div>
    """,
    unsafe_allow_html=True,
)

col1, col2 = st.columns(2, gap="large")
render_upload_slot(col1, 1)
render_upload_slot(col2, 2)

file1_ready = st.session_state.file1_bytes is not None
file2_ready = st.session_state.file2_bytes is not None

st.write("")

proc_col, reset_col = st.columns(2, gap="large")

with proc_col:
    p_l, p_c, p_r = st.columns([1.4, 1, 1.4])
    with p_c:
        process_clicked = st.button(
            "Process Data",
            disabled=not (file1_ready and file2_ready),
            type="primary",
            use_container_width=True,
        )

        # If reconciled, render the high-contrast Download as Excel button right below the Process button as requested
        if st.session_state.reconciled and st.session_state.recon_results is not None:
            st.write("")  # Margin spacer
            excel_data_bytes = st.session_state.recon_results.get("excel_bytes", b"")
            if excel_data_bytes:
                st.download_button(
                    label="📥 Download as Excel",
                    data=excel_data_bytes,
                    file_name="Financial_Reconciliation_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="top_excel_download"
                )

    if not (file1_ready and file2_ready):
        st.markdown(
            "<div style='text-align:center; font-size:0.82rem; color:#4b5563; font-weight: 500;'>Upload both statements to begin processing.</div>",
            unsafe_allow_html=True,
        )

with reset_col:
    if st.session_state.processed:
        r_l, r_c, r_r = st.columns([1.4, 1, 1.4])
        with r_c:
            if st.button("↺ Reset All", type="secondary", use_container_width=True):
                reset_app()
                st.rerun()

# Processing Files Trigger
if process_clicked and file1_ready and file2_ready:
    with st.spinner("Extracting spreadsheets and analyzing row headers..."):
        df1, err1 = load_dataframe(st.session_state.file1_bytes, st.session_state.file1_name)
        df2, err2 = load_dataframe(st.session_state.file2_bytes, st.session_state.file2_name)

    if err1:
        st.error(f"File 1 ({st.session_state.file1_name}): {err1}")
    if err2:
        st.error(f"File 2 ({st.session_state.file2_name}): {err2}")

    if df1 is not None and df2 is not None:
        st.session_state.df1 = df1
        st.session_state.df2 = df2
        st.session_state.processed = True
        st.session_state.reconciled = False
        st.session_state.recon_results = None
        clear_mapping_widget_state()
        st.rerun()
    else:
        st.session_state.processed = False

# Column Mapping Screen
if (
        st.session_state.processed
        and st.session_state.df1 is not None
        and st.session_state.df2 is not None
        and not st.session_state.reconciled
):
    df1 = st.session_state.df1
    df2 = st.session_state.df2
    name1 = st.session_state.file1_name
    name2 = st.session_state.file2_name

    cols1 = list(df1.columns.astype(str))
    cols2 = list(df2.columns.astype(str))
    options1 = [BLANK_OPTION] + cols1
    options2 = [BLANK_OPTION] + cols2

    auto_amt1 = auto_detect_column(cols1, "Amount")
    auto_amt2 = auto_detect_column(cols2, "Amount")

    st.markdown('<div class="section-title">Column Mapping Setup</div>', unsafe_allow_html=True)
    st.caption(
        "Only the **Amount** column is mapped automatically. "
        "All supporting columns default to unmapped unless selected."
    )

    st.markdown('<div class="mapping-card">', unsafe_allow_html=True)

    g_hdr_lbl, g_hdr_1, g_hdr_2 = st.columns([1.2, 2, 2])
    with g_hdr_lbl:
        st.markdown('<div style="font-weight:800; color:#1e3a8a; font-size:1.05rem;">Field Name</div>',
                    unsafe_allow_html=True)
    with g_hdr_1:
        st.markdown(f'<div style="font-weight:800; color:#1e3a8a; font-size:1.05rem;">{name1} Header</div>',
                    unsafe_allow_html=True)
    with g_hdr_2:
        st.markdown(f'<div style="font-weight:800; color:#1e3a8a; font-size:1.05rem;">{name2} Header</div>',
                    unsafe_allow_html=True)
    st.markdown('<div style="margin-top:0.6rem; border-bottom:1px solid #e2e8f0; margin-bottom:0.8rem;"></div>',
                unsafe_allow_html=True)

    selections1, selections2 = {}, {}

    for field in FIXED_FIELDS:
        g_lbl, g_sel1, g_sel2 = st.columns([1.2, 2, 2])
        with g_lbl:
            req_label = "*(Mandatory)*" if field == "Amount" else "(Optional)"
            st.markdown(
                f"<div style='padding-top:0.35rem; font-weight:600;'>{field} <span style='font-size:0.75rem; font-weight:normal; color:#4b5563;'>{req_label}</span></div>",
                unsafe_allow_html=True)

        if field == "Amount":
            idx1 = options1.index(auto_amt1) if auto_amt1 in options1 else 0
            idx2 = options2.index(auto_amt2) if auto_amt2 in options2 else 0
        else:
            idx1 = 0
            idx2 = 0

        with g_sel1:
            selections1[field] = st.selectbox(
                f"sel1_{field}", options1, index=idx1, key=f"map1_{field}", label_visibility="collapsed"
            )
        with g_sel2:
            selections2[field] = st.selectbox(
                f"sel2_{field}", options2, index=idx2, key=f"map2_{field}", label_visibility="collapsed"
            )

        if field == "Amount 2":
            st.markdown(
                "<span style='font-size:0.78rem; color:#2563eb; font-weight:500; display:block; margin-top:-0.5rem; margin-bottom:0.5rem;'>"
                "ℹ️ Select this column only if your payment and invoice is located in different columns"
                "</span>",
                unsafe_allow_html=True
            )
        else:
            st.markdown('<div style="margin:0.4rem 0;"></div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    amount_ok = (
            (selections1["Amount"] != BLANK_OPTION or selections1["Amount 2"] != BLANK_OPTION)
            and (selections2["Amount"] != BLANK_OPTION or selections2["Amount 2"] != BLANK_OPTION)
    )
    can_run = amount_ok

    st.write("")
    run_l, run_c, run_r = st.columns([1.4, 1, 1.4])
    with run_c:
        run_clicked = st.button(
            "Execute Reconciliation",
            disabled=not can_run,
            type="primary",
            use_container_width=True,
        )
    if not can_run:
        st.markdown(
            "<div style='text-align:center; font-size:0.85rem; color:#991b1b; font-weight:700;'>Please map the mandatory 'Amount' columns for both files to proceed.</div>",
            unsafe_allow_html=True,
        )

    if run_clicked and can_run:
        mapping1 = {f: (selections1[f] if selections1[f] != BLANK_OPTION else None) for f in FIXED_FIELDS}
        mapping2 = {f: (selections2[f] if selections2[f] != BLANK_OPTION else None) for f in FIXED_FIELDS}
        with st.spinner("Analyzing rules & classifying balances..."):
            results = run_reconciliation(df1, df2, mapping1, mapping2, name1, name2)
            results["mapping1"] = mapping1
            results["mapping2"] = mapping2

            # Pre-generate the structured Excel file bytes, excluding any unclassified reports
            excel_tables_tracker = build_excel_tables_tracker(results, name1, name2)
            results["excel_bytes"] = generate_excel_report(excel_tables_tracker)

        st.session_state.recon_results = results
        st.session_state.reconciled = True
        st.rerun()

# Displaying Report Screens
if st.session_state.reconciled and st.session_state.recon_results is not None:
    R = st.session_state.recon_results
    name1 = st.session_state.file1_name
    name2 = st.session_state.file2_name

    # ---- 1. Zero-Based Ledger Internal Balance Checklist ----
    st.markdown('<div class="section-title">📊 Zero-Based Internal Balance Check</div>', unsafe_allow_html=True)

    bal1, bal2, bal_diff = R["bal1"], R["bal2"], R["bal_diff"]
    card_class = "balance-card-error" if abs(bal_diff) > 0.01 else "balance-card"
    status_icon = "⚠️" if abs(bal_diff) > 0.01 else "✅"

    balance_text = f"""
    <div class="{card_class}">
        <h4 style="margin-top:0; color:#1e293b;">{status_icon} Zero-Based Net Ledger Summary</h4>
        <p style="margin: 0.5rem 0;">Calculated from absolute 0 based on your strict policy (Payments + Credit Notes + Adjustments - Invoices):</p>
        <p style="margin: 0.2rem 0;"><b>{name1} Net Calculated Balance:</b> ${bal1:,.2f}</p>
        <p style="margin: 0.2rem 0;"><b>{name2} Net Calculated Balance:</b> ${bal2:,.2f}</p>
        <p style="margin: 0.5rem 0 0 0; font-size:1.05rem; font-weight:bold;">
            Net Ledger Variance: <span style="color:{'#dc2626' if abs(bal_diff) > 0.01 else '#16a34a'};">${abs(bal_diff):,.2f}</span>
        </p>
    </div>
    """
    st.markdown(balance_text, unsafe_allow_html=True)

    # Isolated Excluded Folder Report for completely unclassified rows
    unclassified_rows = []
    for u in R["unclassified1"]:
        raw_t = u["raw_type"]
        if is_blank_value(raw_t) or str(raw_t).strip() == "" or str(raw_t).strip().lower() in (
                "nan", "none", "transection type is missing"):
            raw_t_display = "Transection type is missing"
            reason = "Missing transection type"
        else:
            raw_t_display = str(raw_t).strip()
            reason = "Transection type not identified"

        unclassified_rows.append({
            "Source File": name1,
            "Row": u["original_row_num"],
            "Amount": clean_amount_display(u["amount"]),
            "Transection Type": raw_t_display,
            "Reason for Exclusion": reason,
            "Description Context": clean_cell_text(u["orig_desc"]),
            "Reference Context": clean_cell_text(u["orig_ref"])
        })
    for u in R["unclassified2"]:
        raw_t = u["raw_type"]
        if is_blank_value(raw_t) or str(raw_t).strip() == "" or str(raw_t).strip().lower() in (
                "nan", "none", "transection type is missing"):
            raw_t_display = "Transection type is missing"
            reason = "Missing transection type"
        else:
            raw_t_display = str(raw_t).strip()
            reason = "Transection type not identified"

        unclassified_rows.append({
            "Source File": name2,
            "Row": u["original_row_num"],
            "Amount": clean_amount_display(u["amount"]),
            "Transection Type": raw_t_display,
            "Reason for Exclusion": reason,
            "Description Context": clean_cell_text(u["orig_desc"]),
            "Reference Context": clean_cell_text(u["orig_ref"])
        })

    if unclassified_rows:
        with st.expander("📋 Unclassified Transactions Excluded from Balance", expanded=False):
            st.caption(
                "The following transaction records didn't mention or match any target category keywords, so they were excluded from core background balance computations:")
            st.dataframe(pd.DataFrame(unclassified_rows), use_container_width=True, hide_index=True)

    # ---- 2. Data Summary (Initial Transaction Category Summary) ----
    st.markdown('<div class="section-title">🔍 Data Summary</div>', unsafe_allow_html=True)

    categories_list = ["Invoice", "Payment", "Credit Note", "Adjustment", "Unknown"]
    summary_data = []
    payment_diff = R["totals1"]["Payment"] - R["totals2"]["Payment"]
    invoice_diff = R["totals1"]["Invoice"] - R["totals2"]["Invoice"]

    for cat in categories_list:
        total1 = R["totals1"][cat]
        cnt1 = R["counts1"][cat]
        total2 = R["totals2"][cat]
        cnt2 = R["counts2"][cat]
        diff_val = total1 - total2
        clean_cat_lbl = "Unclassified" if cat == "Unknown" else cat

        summary_data.append({
            "Transaction Category": clean_cat_lbl,
            f"{name1} Count": cnt1,
            f"{name1} Total Amount": f"${total1:,.2f}",
            f"{name2} Count": cnt2,
            f"{name2} Total Amount": f"${total2:,.2f}",
            "Count Difference": abs(cnt1 - cnt2),
            "Absolute Difference": f"${abs(diff_val):,.2f}",
            "Status Variance": f"Perfect Align" if abs(diff_val) <= 0.01 else (
                f"${abs(diff_val):,.2f} missing on {name2}" if diff_val > 0 else f"${abs(diff_val):,.2f} missing on {name1}")
        })

    summary_df = pd.DataFrame(summary_data)
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # Immediate Alert notifications
    alerts = []
    if abs(invoice_diff) > 0.01:
        missing_sheet = name2 if invoice_diff > 0 else name1
        alerts.append(f"📌 **Invoice Difference**: ${abs(invoice_diff):,.2f} missing from {missing_sheet}")
    if abs(payment_diff) > 0.01:
        missing_sheet = name2 if payment_diff > 0 else name1
        alerts.append(f"📌 **Payment Difference**: ${abs(payment_diff):,.2f} missing from {missing_sheet}")

    if alerts:
        with st.expander("🚨 Immediate Discrepancies Detected", expanded=True):
            for alert in alerts:
                st.markdown(alert)

    # Extract One-to-Many and Many-to-One Matches out separately
    one_to_many_matches = [m for m in R["advanced_matches"] if "row1" in m]
    many_to_one_matches = [m for m in R["advanced_matches"] if "row2" in m]

    # ---- 3. Reconciliation Results Table Layout ----
    st.markdown('<div class="section-title">📊 Reconciliation Results</div>', unsafe_allow_html=True)

    n_direct = len(R["inter_matches"])
    n_one_to_many = len(one_to_many_matches)
    n_many_to_one = len(many_to_one_matches)
    n_intra = len(R["intra1"]) + len(R["intra2"])
    n_exceptions = len(R["exceptions1"]) + len(R["exceptions2"])
    n_dupes = len(R["dupes1"]) + len(R["dupes2"])

    # Mathematically exact calculation of individual records successfully matched
    matched_indices1 = set()
    matched_indices2 = set()
    for m in R["inter_matches"]:
        matched_indices1.add(m["row1"]["idx"])
        matched_indices2.add(m["row2"]["idx"])
    for m in R["advanced_matches"]:
        if "row1" in m:
            matched_indices1.add(m["row1"]["idx"])
            for r in m["rows2"]:
                matched_indices2.add(r["idx"])
        else:
            matched_indices2.add(m["row2"]["idx"])
            for r in m["rows1"]:
                matched_indices1.add(r["idx"])
    for m in R["intra1"]:
        matched_indices1.add(m["row_a"]["idx"])
        matched_indices1.add(m["row_b"]["idx"])
    for m in R["intra2"]:
        matched_indices2.add(m["row_a"]["idx"])
        matched_indices2.add(m["row_b"]["idx"])

    total_unique_matched = len(matched_indices1) + len(matched_indices2)
    total_analyzed = R["len_rows1"] + R["len_rows2"]
    percentage_matched = (total_unique_matched / total_analyzed * 100) if total_analyzed > 0 else 0.0

    # Build robust results Dataframe aligning with user specified rows
    results_df = pd.DataFrame([
        {"Metric": "Total Records Analyzed in both files", "Result": str(total_analyzed)},
        {"Metric": "Successfully Matched", "Result": str(n_direct)},
        {"Metric": "% of Successfully Matched Transactions", "Result": f"{percentage_matched:.2f}%"},
        {"Metric": "One-to-Many Sum Matched", "Result": str(n_one_to_many)},
        {"Metric": "Many-to-One Sum Matched", "Result": str(n_many_to_one)},
        {"Metric": "Intra-Sheet Offsets", "Result": str(n_intra)},
        {"Metric": "Exceptions Unmatched", "Result": str(n_exceptions)},
        {"Metric": "Possible Duplicates", "Result": str(n_dupes)}
    ])
    st.dataframe(results_df, use_container_width=True, hide_index=True)

    st.write("")

    # Establish precise dynamic mapping flags (Show optional properties ONLY if mapped in both)
    show_date = R["mapping1"].get("Date") is not None and R["mapping2"].get("Date") is not None
    show_ref = R["mapping1"].get("Reference") is not None and R["mapping2"].get("Reference") is not None
    show_desc = R["mapping1"].get("Description") is not None and R["mapping2"].get("Description") is not None

    # ---- 4. Standard Cross-Sheet Matched Transactions (Direct and Cross-Type) ----
    st.markdown('<div class="section-title">✅ Genuine Cross-Sheet Matched Transactions</div>', unsafe_allow_html=True)
    st.caption(
        "Matches containing direct categorizations or adjustments that matched directly across statement files. Every matched row is guaranteed to have amount and row values for both statements.")

    if n_direct:
        matched_rows = []

        for m in R["inter_matches"]:
            r1 = m["row1"]
            r2 = m["row2"]

            matched_props = ["Amount"]
            unmatched_props = []

            if show_date:
                if pd.notna(r1["date"]) and pd.notna(r2["date"]):
                    days = abs((r1["date"] - r2["date"]).days)
                    if days == 0:
                        matched_props.append("Date")
                    else:
                        unmatched_props.append(f"Date ({days} days diff)")
                else:
                    unmatched_props.append("Date blank in record")

            if show_ref:
                ref1 = clean_cell_text(r1["reference"])
                ref2 = clean_cell_text(r2["reference"])
                if ref1 and ref2:
                    if ref1 == ref2:
                        matched_props.append("Reference")
                    else:
                        unmatched_props.append("Reference mismatch")
                else:
                    unmatched_props.append("Reference blank in record")

            status_str = "Matched: " + ", ".join(matched_props)
            if unmatched_props:
                status_str += " | Unmatched: " + ", ".join(unmatched_props)

            # Precise dynamic key matching [Filename]: [Value]
            row_dict = {
                f"{name1}: Row": r1["original_row_num"],
                f"{name2}: Row": r2["original_row_num"],
                f"{name1}: Transaction Type": "" if r1["category"] == "Unknown" else r1["category"],
                f"{name2}: Transaction Type": "" if r2["category"] == "Unknown" else r2["category"],
                f"{name1}: Amount": clean_amount_display(r1["amount"]),
                f"{name2}: Amount": clean_amount_display(r2["amount"]),
            }
            if show_date:
                row_dict[f"{name1}: Date"] = format_date_display(r1["date"])
                row_dict[f"{name2}: Date"] = format_date_display(r2["date"])
            if show_ref:
                row_dict[f"{name1}: Reference"] = ref1
                row_dict[f"{name2}: Reference"] = ref2
            if show_desc:
                row_dict[f"{name1}: Description"] = clean_cell_text(r1["description"])
                row_dict[f"{name2}: Description"] = clean_cell_text(r2["description"])

            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = status_str

            matched_rows.append(row_dict)


        # Payment Matching Prioritization Sort
        def get_match_priority_key(row):
            t1 = str(row.get(f"{name1}: Transaction Type", "")).upper()
            t2 = str(row.get(f"{name2}: Transaction Type", "")).upper()
            mtype = str(row.get("Match Type", "")).upper()
            mstatus = str(row.get("Match Status", "")).upper()

            if "PAYMENT" in t1 and "PAYMENT" in t2:
                return 0
            elif "PAYMENT" in t1 or "PAYMENT" in t2 or "PAYMENT" in mtype or "PAYMENT" in mstatus:
                return 1
            return 2


        matched_rows.sort(key=get_match_priority_key)

        matched_df = pd.DataFrame(matched_rows)
        matched_df = clean_dataframe_strings(matched_df)
        st.dataframe(matched_df, use_container_width=True, hide_index=True)
    else:
        st.info("No cross-statement matches identified.")

    # ---- 4. Separate Table for Intra-Sheet Self-Offsets ----
    st.markdown('<div class="section-title">🔄 Intra-Sheet Self-Offsets (Internal Adjustments)</div>',
                unsafe_allow_html=True)
    st.caption(
        "Self-offsetting transaction entries matched entirely within a single statement file (e.g. Invoices offset internally by Adjustments, or Payments offset by Credit Notes). Since these reconciled internally, they do not require an opposite file amount row.")

    if n_intra:
        intra_rows = []
        for m in R["intra1"]:
            cat_a = m['row_a']['category'] if m['row_a']['category'] != 'Unknown' else 'Unclassified'
            cat_b = m['row_b']['category'] if m['row_b']['category'] != 'Unknown' else 'Unclassified'
            row_dict = {
                "Statement File": name1,
                "Offset Rows": f"Row {m['row_a']['original_row_num']} & Row {m['row_b']['original_row_num']}",
                "Amount": clean_amount_display(m["amount"]),
                "Reconciled Category Types": f"{cat_a} vs {cat_b}",
                "Reconciliation Basis": m["type"],
                "Reconciliation Status": "Matched Internally within File"
            }
            if show_date:
                row_dict[
                    "Dates"] = f"{format_date_display(m['row_a']['date'])}, {format_date_display(m['row_b']['date'])}".strip(
                    ", —")
            if show_ref:
                row_dict[
                    "References"] = f"{clean_cell_text(m['row_a']['reference'])}, {clean_cell_text(m['row_b']['reference'])}".strip(
                    ", ")
            intra_rows.append(row_dict)

        for m in R["intra2"]:
            cat_a = m['row_a']['category'] if m['row_a']['category'] != 'Unknown' else 'Unclassified'
            cat_b = m['row_b']['category'] if m['row_b']['category'] != 'Unknown' else 'Unclassified'
            row_dict = {
                "Statement File": name2,
                "Offset Rows": f"Row {m['row_a']['original_row_num']} & Row {m['row_b']['original_row_num']}",
                "Amount": clean_amount_display(m["amount"]),
                "Reconciled Category Types": f"{cat_a} vs {cat_b}",
                "Reconciliation Basis": m["type"],
                "Reconciliation Status": "Matched Internally within File"
            }
            if show_date:
                row_dict[
                    "Dates"] = f"{format_date_display(m['row_a']['date'])}, {format_date_display(m['row_b']['date'])}".strip(
                    ", —")
            if show_ref:
                row_dict[
                    "References"] = f"{clean_cell_text(m['row_a']['reference'])}, {clean_cell_text(m['row_b']['reference'])}".strip(
                    ", ")
            intra_rows.append(row_dict)

        intra_df = pd.DataFrame(intra_rows)
        intra_df = clean_dataframe_strings(intra_df)
        st.dataframe(intra_df, use_container_width=True, hide_index=True)
    else:
        st.caption("No intra-sheet self-offset entries identified.")

    # ---- 5. One-to-Many Ledger Reconciliation ----
    st.markdown('<div class="section-title">🥞 Split Payments (One-to-Many Ledger Reconciliation)</div>',
                unsafe_allow_html=True)
    st.caption(f"One single transaction from {name1} matched against the combined sum of multiple records in {name2}.")

    if n_one_to_many:
        otm_rows = []
        for m in one_to_many_matches:
            r1 = m["row1"]
            rows2 = m["rows2"]
            r2_idxs = ", ".join(str(r["original_row_num"]) for r in rows2)
            r2_types = ", ".join(r["category"] for r in rows2 if r["category"] != "Unknown")
            r2_refs = ", ".join(clean_cell_text(r["reference"]) for r in rows2 if clean_cell_text(r["reference"]))
            r2_amts = sum(r["amount"] for r in rows2)
            r2_dates = ", ".join(format_date_display(r["date"]) for r in rows2 if pd.notna(r["date"]))
            r2_descs = ", ".join(clean_cell_text(r["description"]) for r in rows2 if clean_cell_text(r["description"]))

            row_dict = {
                f"{name1}: Row": r1["original_row_num"],
                f"{name2}: Row(s)": r2_idxs,
                f"{name1}: Transaction Type": "" if r1["category"] == "Unknown" else r1["category"],
                f"{name2}: Transaction Type(s)": r2_types,
                f"{name1}: Amount": clean_amount_display(r1["amount"]),
                f"{name2}: Amount (Sum)": clean_amount_display(r2_amts),
            }
            if show_date:
                row_dict[f"{name1}: Date"] = format_date_display(r1["date"])
                row_dict[f"{name2}: Date(s)"] = r2_dates
            if show_ref:
                row_dict[f"{name1}: Reference"] = clean_cell_text(r1["reference"])
                row_dict[f"{name2}: Reference(s)"] = r2_refs
            if show_desc:
                row_dict[f"{name1}: Description"] = clean_cell_text(r1["description"])
                row_dict[f"{name2}: Description(s)"] = r2_descs

            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = "Matched: Amount (Sum) | Individual references or dates differ across items"
            otm_rows.append(row_dict)

        otm_df = pd.DataFrame(otm_rows)
        otm_df = clean_dataframe_strings(otm_df)
        st.dataframe(otm_df, use_container_width=True, hide_index=True)
    else:
        st.caption("No One-to-Many matches identified.")

    # ---- 6. Many-to-One Ledger Reconciliation ----
    st.markdown('<div class="section-title">🥞 Combined Payments (Many-to-One Ledger Reconciliation)</div>',
                unsafe_allow_html=True)
    st.caption(f"Multiple transactions from {name1} combined together to match against a single entry in {name2}.")

    if n_many_to_one:
        mto_rows = []
        for m in many_to_one_matches:
            r2 = m["row2"]
            rows1 = m["rows1"]
            r1_idxs = ", ".join(str(r["original_row_num"]) for r in rows1)
            r1_types = ", ".join(r["category"] for r in rows1 if r["category"] != "Unknown")
            r1_refs = ", ".join(clean_cell_text(r["reference"]) for r in rows1 if clean_cell_text(r["reference"]))
            r1_amts = sum(r["amount"] for r in rows1)
            r1_dates = ", ".join(format_date_display(r["date"]) for r in rows1 if pd.notna(r["date"]))
            r1_descs = ", ".join(clean_cell_text(r["description"]) for r in rows1 if clean_cell_text(r["description"]))

            row_dict = {
                f"{name1}: Row(s)": r1_idxs,
                f"{name2}: Row": r2["original_row_num"],
                f"{name1}: Transaction Type(s)": r1_types,
                f"{name2}: Transaction Type": "" if r2["category"] == "Unknown" else r2["category"],
                f"{name1}: Amount (Sum)": clean_amount_display(r1_amts),
                f"{name2}: Amount": clean_amount_display(r2["amount"]),
            }
            if show_date:
                row_dict[f"{name1}: Date(s)"] = r1_dates
                row_dict[f"{name2}: Date"] = format_date_display(r2["date"])
            if show_ref:
                row_dict[f"{name1}: Reference(s)"] = r1_refs
                row_dict[f"{name2}: Reference"] = clean_cell_text(r2["reference"])
            if show_desc:
                row_dict[f"{name1}: Description(s)"] = r1_descs
                row_dict[f"{name2}: Description"] = clean_cell_text(r2["description"])

            row_dict["Match Type"] = m["type"]
            row_dict["Match Score"] = m["score"]
            row_dict["Match Status"] = "Matched: Amount (Sum) | Individual references or dates differ across items"
            mto_rows.append(row_dict)

        mto_df = pd.DataFrame(mto_rows)
        mto_df = clean_dataframe_strings(mto_df)
        st.dataframe(mto_df, use_container_width=True, hide_index=True)
    else:
        st.caption("No Many-to-One matches identified.")

    # ---- 7. Potential Duplicate Entries Table ----
    st.markdown('<div class="section-title">🔁 Potential Duplicate Entries</div>', unsafe_allow_html=True)
    st.caption(
        "Transactions inside a single sheet with identical key attributes (identical amounts, dates, and descriptions) sorted side-by-side for comparison.")

    dup_rows = []
    for dup in R["dupes1"]:
        dup_rows.append({
            "Source Statement": name1,
            "Row": dup["original_row_num"],
            "Amount": clean_cell_text(dup["orig_amount"]) if dup["orig_amount"] else "—",
            "Date": dup["orig_date"] if dup["orig_date"] else "—",
            "Description": dup["orig_desc"] if dup["orig_desc"] else "—",
            "Reference": clean_cell_text(dup["orig_ref"]) if show_ref else "—",
            "Category Type": "" if dup["category"] == "Unknown" else dup["category"],
            "Duplicate Category": dup["exception_type"],
            # Strict sort helpers matching identical criteria
            "_sort_amount": dup["amount"],
            "_sort_date": dup["check_date"],
            "_sort_desc": dup["check_description"].lower().strip()
        })

    for dup in R["dupes2"]:
        dup_rows.append({
            "Source Statement": name2,
            "Row": dup["original_row_num"],
            "Amount": clean_cell_text(dup["orig_amount"]) if dup["orig_amount"] else "—",
            "Date": dup["orig_date"] if dup["orig_date"] else "—",
            "Description": dup["orig_desc"] if dup["orig_desc"] else "—",
            "Reference": clean_cell_text(dup["orig_ref"]) if show_ref else "—",
            "Category Type": "" if dup["category"] == "Unknown" else dup["category"],
            "Duplicate Category": dup["exception_type"],
            # Strict sort helpers matching identical criteria
            "_sort_amount": dup["amount"],
            "_sort_date": dup["check_date"],
            "_sort_desc": dup["check_description"].lower().strip()
        })

    if dup_rows:
        dup_df = pd.DataFrame(dup_rows)
        # Sort sequentially so matched duplicates are perfectly adjacent
        dup_df = dup_df.sort_values(by=["_sort_amount", "_sort_date", "_sort_desc", "Source Statement"],
                                    ascending=[True, True, True, True])
        dup_df = dup_df.drop(columns=["_sort_amount", "_sort_date", "_sort_desc"])
        dup_df = clean_dataframe_strings(dup_df)
        st.dataframe(dup_df, use_container_width=True, hide_index=True)
    else:
        st.caption("No duplicate values detected within either statement.")

    # ---- 7.5 EXCEPTION SUMMARY REPORT (NEW) ----
    if n_exceptions:
        st.markdown('<div class="section-title">📋 Exception Summary</div>', unsafe_allow_html=True)
        st.caption("A high-level count of records missing across your comparison.")

        exc_summary = {
            "Missing Invoices": 0,
            "Missing Payments": 0,
            "Missing Credit Notes": 0,
            "Missing Adjustments": 0
        }
        for exc in R["exceptions1"] + R["exceptions2"]:
            cat = exc["category"]
            if cat == "Invoice":
                exc_summary["Missing Invoices"] += 1
            elif cat == "Payment":
                exc_summary["Missing Payments"] += 1
            elif cat == "Credit Note":
                exc_summary["Missing Credit Notes"] += 1
            elif cat == "Adjustment":
                exc_summary["Missing Adjustments"] += 1

        exc_summary_df = pd.DataFrame([
            {"Issue": "Missing Invoices", "Records": exc_summary["Missing Invoices"]},
            {"Issue": "Missing Payments", "Records": exc_summary["Missing Payments"]},
            {"Issue": "Missing Credit Notes", "Records": exc_summary["Missing Credit Notes"]},
            {"Issue": "Missing Adjustments", "Records": exc_summary["Missing Adjustments"]}
        ])
        st.dataframe(exc_summary_df, use_container_width=True, hide_index=True)

    # ---- 8. Exceptions Report (Updated with Source Context) ----
    st.markdown('<div class="section-title">⚠️ Exceptions & Unmatched Report</div>', unsafe_allow_html=True)
    st.caption(
        "Transactions that failed to match on amounts, are completely omitted from statements, or have no defined amounts.")

    unmatched_rows = []

    for exc in R["exceptions1"]:
        row_dict = {
            "Source Statement": name1,
            "Row": exc["original_row_num"],
            "Amount": clean_cell_text(exc["orig_amount"]) if exc["orig_amount"] else "—",
        }
        if show_date:
            row_dict["Date"] = clean_cell_text(exc["orig_date"]) if exc["orig_date"] else "—"
        if show_ref:
            row_dict["Reference"] = clean_cell_text(exc["orig_ref"]) if exc["orig_ref"] else "—"
        row_dict.update({
            "Category Type": "" if exc["category"] == "Unknown" else exc["category"],
            "Description Context": clean_cell_text(exc["orig_desc"]) if exc["orig_desc"] else "—",
            "Exception Category": exc["exception_type"]
        })
        unmatched_rows.append(row_dict)

    for exc in R["exceptions2"]:
        row_dict = {
            "Source Statement": name2,
            "Row": exc["original_row_num"],
            "Amount": clean_cell_text(exc["orig_amount"]) if exc["orig_amount"] else "—",
        }
        if show_date:
            row_dict["Date"] = clean_cell_text(exc["orig_date"]) if exc["orig_date"] else "—"
        if show_ref:
            row_dict["Reference"] = clean_cell_text(exc["orig_ref"]) if exc["orig_ref"] else "—"
        row_dict.update({
            "Category Type": "" if exc["category"] == "Unknown" else exc["category"],
            "Description Context": clean_cell_text(exc["orig_desc"]) if exc["orig_desc"] else "—",
            "Exception Category": exc["exception_type"]
        })
        unmatched_rows.append(row_dict)

    if unmatched_rows:
        exceptions_df = pd.DataFrame(unmatched_rows)
        # Create helper numeric sorting column to arrange exceptions alphabetically or numerically
        exceptions_df["_sort_amount"] = exceptions_df["Amount"].apply(lambda x: parse_amount(x))
        exceptions_df = exceptions_df.sort_values(by=["_sort_amount", "Source Statement"], ascending=[True, True])
        exceptions_df = exceptions_df.drop(columns=["_sort_amount"])
        exceptions_df = clean_dataframe_strings(exceptions_df)
        st.dataframe(exceptions_df, use_container_width=True, hide_index=True)
    else:
        st.success("Perfect reconciliation! All entries have successfully matched.")
